"""ANFAVEA vehicle production/sales source — Brazil automotive industry data.

Downloads and parses ANFAVEA monthly Excel files containing vehicle
production, licensing, and export data for Brazil.

Uses BrowserManager to navigate to the ANFAVEA editions page and locate
the latest Excel download link.

URL: https://anfavea.com.br/site/edicoes-em-excel/
"""

from __future__ import annotations

import io
import logging
import re
from datetime import datetime

from pydantic import BaseModel

from openquery.exceptions import SourceError
from openquery.models.br.anfavea import AnfaveaResult, AnfaveaSegment
from openquery.sources import register
from openquery.sources.base import BaseSource, DocumentType, QueryInput, SourceMeta

logger = logging.getLogger(__name__)

ANFAVEA_URL = "https://anfavea.com.br/site/edicoes-em-excel/"

# Segment row keywords to look for in the Excel sheet
SEGMENT_KEYWORDS = {
    "automobiles": ["autom", "automóveis", "automoveis"],
    "light_commercial": ["veículos comerciais leves", "comerciais leves", "veiculo comercial"],
    "trucks": ["caminh", "caminhoes", "caminhões"],
    "buses": ["ônibus", "onibus"],
}


@register
class AnfaveaSource(BaseSource):
    """Query ANFAVEA Brazil vehicle production/licensing/export statistics."""

    def __init__(self, timeout: float = 60.0, headless: bool = True) -> None:
        self._timeout = timeout
        self._headless = headless

    def meta(self) -> SourceMeta:
        return SourceMeta(
            name="br.anfavea",
            display_name="ANFAVEA — Produção e Emplacamentos de Veículos",
            description=(
                "Brazil ANFAVEA monthly vehicle statistics: production, licensing, "
                "and exports by segment (automobiles, light commercial, trucks, buses)"
            ),
            country="BR",
            url=ANFAVEA_URL,
            supported_inputs=[DocumentType.CUSTOM],
            requires_captcha=False,
            requires_browser=True,
            rate_limit_rpm=5,
        )

    def query(self, input: QueryInput) -> BaseModel:
        year = input.extra.get("year", "").strip()
        month = input.extra.get("month", "").strip()
        return self._fetch(year, month)

    def _fetch(self, year: str, month: str) -> AnfaveaResult:
        try:
            import httpx

            from openquery.core.browser import BrowserManager

            logger.info(
                "Fetching ANFAVEA Excel: year=%s month=%s", year or "latest", month or "latest"
            )

            browser = BrowserManager(headless=self._headless, timeout=self._timeout)
            excel_url = None

            with browser.sync_context() as ctx:
                page = ctx.new_page()
                page.goto(
                    ANFAVEA_URL,
                    wait_until="domcontentloaded",
                    timeout=int(self._timeout * 1000),
                )

                # Find Excel download links (.xlsx or .xls)
                links = page.query_selector_all("a[href*='.xls']")
                candidates: list[str] = []
                for link in links:
                    href = link.get_attribute("href") or ""
                    if href:
                        if not href.startswith("http"):
                            href = "https://anfavea.com.br" + href
                        candidates.append(href)

            if not candidates:
                raise SourceError("br.anfavea", "No Excel links found on ANFAVEA page")

            # Filter by year if provided
            if year:
                year_candidates = [c for c in candidates if year in c]
                if year_candidates:
                    candidates = year_candidates

            # Use the first (most recent) candidate
            excel_url = candidates[0]
            logger.info("Downloading ANFAVEA Excel: %s", excel_url)

            headers = {
                "User-Agent": "Mozilla/5.0 (compatible; OpenQuery/0.9.0)",
                "Accept": "application/octet-stream, application/vnd.ms-excel, */*",
                "Referer": ANFAVEA_URL,
            }
            with httpx.Client(
                timeout=self._timeout, headers=headers, follow_redirects=True
            ) as client:
                resp = client.get(excel_url)
                resp.raise_for_status()
                excel_bytes = resp.content

            return self._parse_excel(excel_bytes, excel_url, year, month)

        except httpx.HTTPStatusError as e:
            raise SourceError(
                "br.anfavea", f"HTTP {e.response.status_code} downloading Excel"
            ) from e
        except httpx.RequestError as e:
            raise SourceError("br.anfavea", f"Request failed: {e}") from e
        except SourceError:
            raise
        except Exception as e:
            raise SourceError("br.anfavea", f"Query failed: {e}") from e

    def _parse_excel(
        self, excel_bytes: bytes, excel_url: str, year: str, month: str
    ) -> AnfaveaResult:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        except Exception:
            try:
                import xlrd

                wb_xlrd = xlrd.open_workbook(file_contents=excel_bytes)
                return self._parse_xlrd(wb_xlrd, excel_url, year, month)
            except Exception as e:
                raise SourceError("br.anfavea", f"Cannot parse Excel file: {e}") from e

        # Try to find the production/licensing sheet
        sheet_names = wb.sheetnames
        target_sheet = wb.active
        for name in sheet_names:
            nl = name.lower()
            if any(k in nl for k in ["prod", "emplacamento", "export", "total"]):
                target_sheet = wb[name]
                break

        segments: list[AnfaveaSegment] = []
        total_production = 0
        total_licensing = 0
        total_exports = 0

        rows = list(target_sheet.iter_rows(values_only=True))

        # Determine period from URL or year/month params
        period = _extract_period(excel_url, year, month)

        # Scan rows for segment data
        for row in rows:
            if not row or row[0] is None:
                continue
            cell0 = str(row[0]).lower().strip()

            for seg_name, keywords in SEGMENT_KEYWORDS.items():
                if any(kw in cell0 for kw in keywords):
                    prod = _safe_int(row[1] if len(row) > 1 else None)
                    lic = _safe_int(row[2] if len(row) > 2 else None)
                    exp = _safe_int(row[3] if len(row) > 3 else None)
                    segments.append(
                        AnfaveaSegment(
                            segment=seg_name,
                            production=prod,
                            licensing=lic,
                            exports=exp,
                        )
                    )
                    total_production += prod
                    total_licensing += lic
                    total_exports += exp
                    break

        return AnfaveaResult(
            queried_at=datetime.now(),
            period=period,
            total_production=total_production,
            total_licensing=total_licensing,
            total_exports=total_exports,
            segments=segments,
        )

    def _parse_xlrd(self, wb, excel_url: str, year: str, month: str) -> AnfaveaResult:
        """Fallback parser for .xls files using xlrd."""
        sheet = wb.sheet_by_index(0)
        segments: list[AnfaveaSegment] = []
        total_production = 0
        total_licensing = 0
        total_exports = 0
        period = _extract_period(excel_url, year, month)

        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            if not row or not row[0]:
                continue
            cell0 = str(row[0]).lower().strip()

            for seg_name, keywords in SEGMENT_KEYWORDS.items():
                if any(kw in cell0 for kw in keywords):
                    prod = _safe_int(row[1] if len(row) > 1 else None)
                    lic = _safe_int(row[2] if len(row) > 2 else None)
                    exp = _safe_int(row[3] if len(row) > 3 else None)
                    segments.append(
                        AnfaveaSegment(
                            segment=seg_name,
                            production=prod,
                            licensing=lic,
                            exports=exp,
                        )
                    )
                    total_production += prod
                    total_licensing += lic
                    total_exports += exp
                    break

        return AnfaveaResult(
            queried_at=datetime.now(),
            period=period,
            total_production=total_production,
            total_licensing=total_licensing,
            total_exports=total_exports,
            segments=segments,
        )


def _safe_int(val: object) -> int:
    """Safely convert a cell value to int."""
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(",", "").replace(".", "").strip() or "0"))
    except (ValueError, TypeError):
        return 0


def _extract_period(url: str, year: str, month: str) -> str:
    """Extract period string from URL or params."""
    if year and month:
        return f"{year}-{month.zfill(2)}"
    if year:
        return year
    # Try to extract year from URL
    match = re.search(r"(20\d{2})", url)
    if match:
        return match.group(1)
    return datetime.now().strftime("%Y")
